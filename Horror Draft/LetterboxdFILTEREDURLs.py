from openpyxl import load_workbook
from openpyxl import Workbook

# Load the source workbook
source_workbook = load_workbook('C:\\Users\\magic\\OneDrive\\Desktop\\School Stuff\\Python\\href_elements.xlsx')
source_sheet = source_workbook.active

# Create a new workbook to store the filtered data
target_workbook = Workbook()
target_sheet = target_workbook.active

# Iterate through each row in column A of the source workbook
for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=1, values_only=True):
    cell_value = row[0]  # Get the value of the cell in column A
    
    # Check if the cell value contains "/film/"
    if cell_value and "/film/" in cell_value:
        # Write the cell value to the target workbook
        target_sheet.append([cell_value])

# Save the target workbook
target_workbook.save('filtered_href_elements.xlsx')
